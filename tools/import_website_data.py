"""
Import Website Tire Data into Toko Ban Database
=================================================
Extracts 13 brands, 1872+ tire products from the website and
saves directly to products.xlsx using pandas.

Usage: PYTHONIOENCODING=utf-8 python tools/import_website_data.py
"""

import re
import json
import sys
from pathlib import Path
from datetime import datetime

WEBSITE_DIR = Path("C:/Users/ASUS/AI Apps/(1)Toko Ban Murah Anugerah Website")
DATABASE_DIR = Path(__file__).parent.parent


def js_to_json(c):
    """Convert JS var brands = {...} to valid JSON."""
    c = c.strip()
    if c.startswith('var brands = '):
        c = c[len('var brands = '):]
    if c.endswith('};'):
        c = c[:-1]
    if c.endswith(';'):
        c = c[:-1]
    c = re.sub(r'//.*?(\n|$)', '\n', c)
    c = re.sub(r'/\*.*?\*/', '', c, flags=re.DOTALL)
    result = []
    i = 0
    while i < len(c):
        if c[i] == "'":
            j = i + 1
            while j < len(c):
                if c[j] == '\\' and j + 1 < len(c):
                    j += 2
                    continue
                if c[j] == "'":
                    break
                j += 1
            inner = c[i+1:j].replace('"', '\\"')
            result.append('"' + inner + '"')
            i = j + 1
        else:
            result.append(c[i])
            i += 1
    c = ''.join(result)
    c = re.sub(r'(\s+)(\w[\w\d]*)(\s*):(\s*)', r'\1"\2"\3:\4', c)
    c = re.sub(r',(\s*[}\]])', r'\1', c)
    return c


def extract_ring(sz):
    m = re.search(r'[RZ](\d{2})', sz.replace(' ', ''), re.I)
    return m.group(1) if m else ''


def estimate_price(brand, ring, cat):
    r = int(ring) if ring.isdigit() else 15
    base = {12:300000,13:350000,14:450000,15:550000,16:650000,
            17:850000,18:1000000,19:1300000,20:1700000,
            21:2200000,22:2800000,23:3500000,24:4000000}
    p = base.get(r, 500000)
    if cat == 'Ban SUV':
        p = int(p * 1.15)
    if brand in {'Bridgestone','Michelin','Yokohama','Goodyear'}:
        p = int(p * 1.3)
    if brand in {'Accelera','Delium','Blackhawk','Swallow','Sailun','GT Radial'}:
        p = int(p * 0.85)
    return max(round(p / 10000) * 10000, 300000)


def main():
    print("=" * 60)
    print("IMPORT WEBSITE TIRE DATA TO DATABASE")
    print("=" * 60)

    brand_js = WEBSITE_DIR / "js" / "brands-data.js"
    if not brand_js.exists():
        print(f"ERROR: {brand_js} not found")
        return

    print(f"\nReading brand data...")
    with open(str(brand_js), "r", encoding="utf-8") as f:
        content = f.read()

    json_str = js_to_json(content)
    brands = json.loads(json_str)
    print(f"Loaded {len(brands)} brands")

    products = []
    seen = set()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    for brand_key, brand_data in brands.items():
        name = brand_data.get("name", brand_key.title())
        for model in brand_data.get("models", []):
            mname = model.get("name", "Unknown")
            for subcat in model.get("subCategories", []):
                for entry in subcat.get("entries", []):
                    sz = entry.get("size", "").strip()
                    if not sz:
                        continue
                    ring = extract_ring(sz)
                    if not ring:
                        continue
                    key = f"{name}|{mname}|{sz}"
                    if key in seen:
                        continue
                    seen.add(key)

                    # Determine category
                    s = sz.upper().replace(' ', '')
                    cat = "Ban Mobil"
                    if "LT" in s or ("C" in s and int(ring) >= 14):
                        cat = "Ban Truk"
                    elif any(w in s for w in ["265","275","285","295","305","315","325","33X","35X","37X"]):
                        cat = "Ban SUV"
                    elif int(ring) >= 19:
                        cat = "Ban SUV"

                    price = estimate_price(name, ring, cat)
                    stock = 10 if cat == "Ban Mobil" else 8

                    products.append({
                        "ID": f"TB-{len(products)+1:04d}",
                        "Merek": name,
                        "Tipe": mname,
                        "Ukuran": sz,
                        "Ring": ring,
                        "Harga Beli": int(price * 0.7),
                        "Harga Jual": price,
                        "Stok": stock,
                        "Stok Minimum": 5,
                        "Supplier": "",
                        "Kategori": cat,
                        "Lokasi Rak": f"Rak-{name[:3].upper()}-{ring}",
                        "Terakhir Update": now,
                    })

    print(f"Extracted {len(products)} unique products")

    # Count per brand
    bcount = {}
    for p in products:
        bcount[p["Merek"]] = bcount.get(p["Merek"], 0) + 1
    for b, c in sorted(bcount.items(), key=lambda x: -x[1]):
        print(f"  - {b}: {c}")

    # Save using openpyxl directly (most reliable for xlsx)
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    
    data_dir = DATABASE_DIR / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    filepath = data_dir / "products.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Headers
    headers = ["ID","Merek","Tipe","Ukuran","Ring","Harga Beli","Harga Jual","Stok","Stok Minimum","Supplier","Kategori","Lokasi Rak","Terakhir Update"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    # Data rows
    for row_idx, prod in enumerate(products, 2):
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=prod.get(h, ""))
    
    wb.save(str(filepath))
    print(f"\nSaved {len(products)} products to {filepath}")
    print(f"File size: {filepath.stat().st_size:,} bytes")
    
    # Verify with InventoryManager
    sys.path.insert(0, str(DATABASE_DIR))
    from tools.inventory import InventoryManager
    inv = InventoryManager(str(filepath))
    s = inv.get_summary()
    print(f"Verified: {s['total_items']} items, {s['unique_brands']} brands, {s['total_stock']} total stock")


if __name__ == "__main__":
    main()
