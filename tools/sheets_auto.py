"""
Google Sheets / Excel Automation - Toko Ban Murah Anugerah
===========================================================
Automate Excel reports, merge data from multiple sources,
create pivot tables, charts, and email reports automatically.
"""

import csv
import json
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional

try:
    import pandas as pd
    _has_pandas = True
except ImportError:
    _has_pandas = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    _has_openpyxl = True
except ImportError:
    _has_openpyxl = False


class SheetsAutomation:
    """Automate Excel/Sheets reporting and data processing."""

    def __init__(self, output_dir: str = "../output/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ─── Create Sales Report ─────────────────────────────────────────────
    def create_sales_report(self, sales_data: list, 
                            filename: Optional[str] = None) -> Optional[str]:
        """Generate a formatted Excel sales report with charts."""
        if not _has_openpyxl or not _has_pandas:
            return None
        
        wb = openpyxl.Workbook()
        
        # ── Sheet 1: Summary ──
        ws1 = wb.active
        ws1.title = "Ringkasan"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003366", end_color="003366", 
                                  fill_type="solid")
        
        # Title
        ws1.cell(row=1, column=1, value="LAPORAN PENJUALAN").font = Font(
            bold=True, size=16, color="003366")
        ws1.merge_cells("A1:E1")
        
        ws1.cell(row=2, column=1, 
                 value=f"Toko Ban Murah Anugerah - {datetime.now().strftime('%B %Y')}").font = Font(
            size=11, color="666666")
        
        # Calculate summary
        total_revenue = sum(s.get("Total", 0) for s in sales_data)
        total_transactions = len(sales_data)
        total_profit = sum(s.get("Keuntungan", 0) for s in sales_data)
        avg_transaction = total_revenue / total_transactions if total_transactions else 0
        
        # Summary table
        headers = ["Metrik", "Nilai"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=5, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        summary_data = [
            ("Total Pendapatan", f"Rp {total_revenue:,.0f}"),
            ("Total Transaksi", total_transactions),
            ("Total Keuntungan", f"Rp {total_profit:,.0f}"),
            ("Rata-rata Transaksi", f"Rp {avg_transaction:,.0f}"),
        ]
        for i, (metric, value) in enumerate(summary_data):
            ws1.cell(row=6+i, column=1, value=metric)
            ws1.cell(row=6+i, column=2, value=value)
        
        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 20

        # ── Sheet 2: Detail Transaksi ──
        ws2 = wb.create_sheet("Detail Transaksi")
        
        if sales_data:
            df = pd.DataFrame(sales_data)
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
            
            # Add auto-filter
            if df.shape[1] > 0:
                ws2.auto_filter.ref = f"A1:{chr(64 + df.shape[1])}{len(df) + 1}"

        # ── Sheet 3: Produk Terlaris ──
        ws3 = wb.create_sheet("Produk Terlaris")
        
        product_sales = {}
        for s in sales_data:
            prod = s.get("Produk", "Unknown")
            product_sales[prod] = product_sales.get(prod, 0) + s.get("Jumlah", 1)
        
        sorted_products = sorted(product_sales.items(), 
                                 key=lambda x: x[1], reverse=True)
        
        ws3.cell(row=1, column=1, value="Produk").font = header_font
        ws3.cell(row=1, column=1).fill = header_fill
        ws3.cell(row=1, column=2, value="Terjual").font = header_font
        ws3.cell(row=1, column=2).fill = header_fill
        
        for i, (prod, qty) in enumerate(sorted_products, 2):
            ws3.cell(row=i, column=1, value=prod)
            ws3.cell(row=i, column=2, value=qty)
        
        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 15
        
        # Bar chart
        if len(sorted_products) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.title = "10 Produk Terlaris"
            chart.y_axis.title = "Jumlah Terjual"
            
            data_ref = Reference(ws3, min_col=2, min_row=1, 
                                 max_row=min(11, len(sorted_products) + 1))
            cats_ref = Reference(ws3, min_col=1, min_row=2,
                                 max_row=min(11, len(sorted_products) + 1))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            
            ws3.add_chart(chart, "D2")

        # Save
        if not filename:
            filename = f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        filepath = self.output_dir / filename
        wb.save(str(filepath))
        return str(filepath)

    # ─── Merge Product & Inventory Data ──────────────────────────────────
    def create_catalog_excel(self, products: list, 
                             filename: Optional[str] = None) -> Optional[str]:
        """Create a professional product catalog Excel file."""
        if not _has_openpyxl:
            return None
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Katalog Ban"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003366", end_color="003366", 
                                  fill_type="solid")
        alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", 
                               fill_type="solid")
        
        # Title
        ws.cell(row=1, column=1, value="KATALOG BAN - Toko Ban Murah Anugerah").font = Font(
            bold=True, size=14, color="003366")
        ws.merge_cells("A1:F1")
        
        # Headers
        headers = ["Merek", "Tipe", "Ukuran", "Ring", "Harga Jual", "Stok"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for i, prod in enumerate(products):
            row = i + 4
            values = [
                prod.get("Merek", ""),
                prod.get("Tipe", ""),
                prod.get("Ukuran", ""),
                str(prod.get("Ring", "")),
                prod.get("Harga Jual", 0),
                prod.get("Stok", 0)
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if i % 2 == 1:
                    cell.fill = alt_fill
                if col == 5:  # Price column
                    cell.number_format = 'Rp #,##0'
        
        # Column widths
        widths = [20, 25, 18, 8, 15, 10]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i)].width = w
        
        if not filename:
            filename = f"catalog_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        filepath = self.output_dir.parent / "price_lists" / filename
        filepath.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(filepath))
        return str(filepath)

    # ─── Export to CSV ───────────────────────────────────────────────────
    def export_to_csv(self, data: list, fieldnames: list,
                      filename: str) -> str:
        """Export data to CSV file."""
        filepath = self.output_dir / filename
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        return str(filepath)

    # ─── Generate Daily Summary ──────────────────────────────────────────
    def daily_summary_text(self, sales_data: list, inventory_summary: dict,
                          orders_summary: dict) -> str:
        """Generate a daily business summary text for WhatsApp."""
        today = datetime.now().strftime("%d-%m-%Y")
        
        total_revenue = sum(s.get("Total", 0) for s in sales_data)
        
        lines = [
            f"📊 *LAPORAN HARIAN - {today}*",
            f"🏪 Toko Ban Murah Anugerah",
            "━" * 30,
            "",
            "💰 *PENJUALAN:*",
            f"  Transaksi: {len(sales_data)}",
            f"  Pendapatan: Rp {total_revenue:,.0f}",
            "",
            "📦 *INVENTORY:*",
            f"  Total Produk: {inventory_summary.get('total_items', 0)}",
            f"  Stok Menipis: {inventory_summary.get('low_stock_count', 0)}",
            f"  Stok Habis: {inventory_summary.get('out_of_stock_count', 0)}",
            "",
            "📋 *PESANAN:*",
            f"  Pending: {orders_summary.get('pending', 0)}",
            f"  Baru: {orders_summary.get('total', 0)}",
            "",
            "✅ Sukses hari ini! 💪"
        ]
        
        return "\n".join(lines)
