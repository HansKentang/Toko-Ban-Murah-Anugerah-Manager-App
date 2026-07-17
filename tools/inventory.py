"""
Inventory & Stock Manager - Toko Ban Murah Anugerah
====================================================
Features:
- Track tire stock by brand, size, type
- Low stock alerts
- Supplier management
- Stock history / movement tracking
- Import/Export from Excel
"""

import json
import csv
from pathlib import Path
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    import pandas as pd
    _has_pandas = True
except ImportError:
    _has_pandas = False


PRODUCT_FIELDS = [
    "ID", "Merek", "Tipe", "Ukuran", "Ring", "Harga Beli", 
    "Harga Jual", "Stok", "Stok Minimum", "Supplier", 
    "Kategori", "Lokasi Rak", "Terakhir Update"
]

CATEGORIES = ["Ban Mobil", "Ban Motor", "Ban Truk", "Ban SUV", "Ban Vintage"]
TIRE_BRANDS = ["Bridgestone", "Michelin", "Dunlop", "GT Radial", "Achilles", 
               "Continental", "Hankook", "Goodyear", "Pirelli", "Yokohama",
               "IRC", "Swallow", "FDR", "Corsa", "Maxxis"]

TIRE_SIZES_CAR = [
    "155/65R13", "155/70R13", "165/70R13", "175/70R13",
    "175/65R14", "185/70R14", "185/65R14", "195/65R14",
    "195/55R15", "195/60R15", "205/65R15", "205/55R16",
    "215/55R17", "225/45R17", "225/55R18"
]


class InventoryManager:
    """Complete inventory management for tire shop."""

    def __init__(self, data_path: str = "../data/products.xlsx"):
        self.data_path = Path(data_path)
        self.products = []
        self._load_products()

    def _load_products(self):
        """Load products from Excel or create template if not exists."""
        if self.data_path.exists():
            try:
                if _has_pandas:
                    df = pd.read_excel(self.data_path)
                    self.products = df.to_dict("records")
                else:
                    wb = openpyxl.load_workbook(self.data_path)
                    ws = wb.active
                    headers = [cell.value for cell in ws[1]]
                    self.products = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(cell is not None for cell in row):
                            self.products.append(dict(zip(headers, row)))
                return
            except Exception:
                pass
        
        # Create template
        self.products = []
        self._save_products()

    def _save_products(self):
        """Save products to Excel file."""
        self.data_path.parent.mkdir(parents=True, exist_ok=True)
        
        if _has_pandas:
            df = pd.DataFrame(self.products)
            if not df.empty:
                df.to_excel(self.data_path, index=False)
            else:
                pd.DataFrame(columns=PRODUCT_FIELDS).to_excel(self.data_path, index=False)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(PRODUCT_FIELDS)
            for prod in self.products:
                ws.append([prod.get(f, "") for f in PRODUCT_FIELDS])
            wb.save(self.data_path)

    # ─── CRUD Operations ────────────────────────────────────────────────
    def add_product(self, product: dict) -> dict:
        """Add a new product to inventory."""
        product["ID"] = product.get("ID", f"TB-{len(self.products) + 1:04d}")
        product["Stok"] = int(product.get("Stok", 0))
        product["Stok Minimum"] = int(product.get("Stok Minimum", 5))
        product["Harga Beli"] = float(product.get("Harga Beli", 0))
        product["Harga Jual"] = float(product.get("Harga Jual", 0))
        product["Terakhir Update"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        self.products.append(product)
        self._save_products()
        return product

    def update_product(self, product_id: str, updates: dict) -> Optional[dict]:
        """Update a product by ID."""
        for i, prod in enumerate(self.products):
            if prod.get("ID") == product_id:
                updates["Terakhir Update"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                self.products[i].update(updates)
                self._save_products()
                return self.products[i]
        return None

    def delete_product(self, product_id: str) -> bool:
        """Delete a product by ID."""
        for i, prod in enumerate(self.products):
            if prod.get("ID") == product_id:
                self.products.pop(i)
                self._save_products()
                return True
        return False

    def get_product(self, product_id: str) -> Optional[dict]:
        """Get a single product by ID."""
        for prod in self.products:
            if prod.get("ID") == product_id:
                return prod
        return None

    def search_products(self, query: str) -> list:
        """Search products by brand, type, or size."""
        query = query.lower()
        results = []
        for prod in self.products:
            if (query in str(prod.get("Merek", "")).lower() or
                query in str(prod.get("Tipe", "")).lower() or
                query in str(prod.get("Ukuran", "")).lower()):
                results.append(prod)
        return results

    # ─── Stock Operations ───────────────────────────────────────────────
    def add_stock(self, product_id: str, quantity: int) -> Optional[dict]:
        """Add stock to a product."""
        return self.update_product(product_id, {
            "Stok": self.get_product(product_id)["Stok"] + quantity
        })

    def remove_stock(self, product_id: str, quantity: int) -> Optional[dict]:
        """Remove stock from a product."""
        prod = self.get_product(product_id)
        if not prod:
            return None
        new_stock = max(0, prod["Stok"] - quantity)
        return self.update_product(product_id, {"Stok": new_stock})

    def get_low_stock_products(self, threshold: Optional[int] = None) -> list:
        """Get products with stock below threshold."""
        if threshold is None:
            threshold = 5
        return [p for p in self.products if p.get("Stok", 0) <= threshold]

    def get_out_of_stock(self) -> list:
        """Get products with zero stock."""
        return [p for p in self.products if p.get("Stok", 0) == 0]

    # ─── Reports ────────────────────────────────────────────────────────
    def get_summary(self) -> dict:
        """Get inventory summary statistics."""
        total_items = len(self.products)
        total_stock = sum(p.get("Stok", 0) for p in self.products)
        total_value = sum(p.get("Harga Beli", 0) * p.get("Stok", 0) 
                         for p in self.products)
        low_stock = len(self.get_low_stock_products())
        out_of_stock = len(self.get_out_of_stock())
        
        # Group by brand
        brands = {}
        for p in self.products:
            brand = p.get("Merek", "Unknown")
            brands[brand] = brands.get(brand, 0) + 1
        
        return {
            "total_items": total_items,
            "total_stock": total_stock,
            "total_value": total_value,
            "low_stock_count": low_stock,
            "out_of_stock_count": out_of_stock,
            "unique_brands": len(brands),
            "brands": brands
        }

    def export_csv(self, filepath: str):
        """Export inventory to CSV."""
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=PRODUCT_FIELDS)
            writer.writeheader()
            writer.writerows(self.products)

    # ─── Quick Add Sample Data ─────────────────────────────────────────
    def add_sample_data(self):
        """Add sample tire products for demonstration."""
        samples = [
            {"Merek": "Bridgestone", "Tipe": "Turanza", "Ukuran": "195/65R15", 
             "Ring": "15", "Harga Beli": 450000, "Harga Jual": 650000, 
             "Stok": 12, "Stok Minimum": 5, "Kategori": "Ban Mobil"},
            {"Merek": "Michelin", "Tipe": "Energy XM2", "Ukuran": "185/65R14", 
             "Ring": "14", "Harga Beli": 520000, "Harga Jual": 750000, 
             "Stok": 8, "Stok Minimum": 3, "Kategori": "Ban Mobil"},
            {"Merek": "GT Radial", "Tipe": "Champiro", "Ukuran": "175/70R13", 
             "Ring": "13", "Harga Beli": 280000, "Harga Jual": 400000, 
             "Stok": 20, "Stok Minimum": 10, "Kategori": "Ban Mobil"},
            {"Merek": "Dunlop", "Tipe": "SP Sport", "Ukuran": "215/55R17", 
             "Ring": "17", "Harga Beli": 680000, "Harga Jual": 950000, 
             "Stok": 4, "Stok Minimum": 3, "Kategori": "Ban Mobil"},
            {"Merek": "IRC", "Tipe": "NR77", "Ukuran": "80/90-17", 
             "Ring": "17", "Harga Beli": 95000, "Harga Jual": 150000, 
             "Stok": 25, "Stok Minimum": 10, "Kategori": "Ban Motor"},
            {"Merek": "Michelin", "Tipe": "Pilot Street", "Ukuran": "90/80-17", 
             "Ring": "17", "Harga Beli": 180000, "Harga Jual": 280000, 
             "Stok": 10, "Stok Minimum": 5, "Kategori": "Ban Motor"},
            {"Merek": "Achilles", "Tipe": "ATR", "Ukuran": "205/65R15", 
             "Ring": "15", "Harga Beli": 350000, "Harga Jual": 500000, 
             "Stok": 15, "Stok Minimum": 8, "Kategori": "Ban Mobil"},
        ]
        for s in samples:
            self.add_product(s)
        return len(samples)
